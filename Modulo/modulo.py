import os
import platform
import subprocess
import math
import requests             
from datetime import datetime

import openpyxl
try:

    #Crie um novo arquivo .xlsx
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Lista Resumida"  
    sheet2 = workbook.create_sheet(title="Lista Completa")
    sheet3 = workbook.create_sheet(title="Hardware")

    # Construa o caminho para a pasta de documentos e cria a pasta do programa
    doc = os.path.join(os.path.expanduser("~"), "Documents")
    path = os.path.join(doc, "AferidorDesktop")

    # Crie a pasta se ela não existir
    if not os.path.exists(path):
        os.makedirs(path)

    planilha_path = os.path.join(path, "LOG_AferidorDektop.xlsx")
    txt_path = os.path.join(path, "LOG_AferidorDesktop.txt")

    output_file = txt_path

    '''
    Links Uteis
    https://learn.microsoft.com/en-us/windows/win32/cimwin32prov/computer-system-hardware-classes
    https://learn.microsoft.com/pt-br/windows/win32/cimwin32prov/computer-system-hardware-classes
    https://learn.microsoft.com/pt-br/windows/win32/wmisdk/wmi-tasks--computer-hardware
    '''

    # Dicionário que relaciona os números aos valores correspondentes
    modelos_hardware = {
        1: 'Other',
        2: 'Unknown',
        3: 'Desktop',
        4: 'Low Profile Desktop',
        5: 'Pizza Box',
        6: 'Mini Tower',
        7: 'Tower',
        8: 'Portable',
        9: 'Laptop',
        10: 'Notebook',
        11: 'Hand Held',
        12: 'Docking Station',
        13: 'All in One',
        14: 'Sub Notebook',
        15: 'Space-Saving',
        16: 'Lunch Box',
        17: 'Main System Chassis',
        18: 'Expansion Chassis',
        19: 'SubChassis',
        20: 'Bus Expansion Chassis',
        21: 'Peripheral Chassis',
        22: 'Storage Chassis',
        23: 'Rack Mount Chassis',
        24: 'Sealed-Case PC',
        30: 'Tablet',
        31: 'Convertible',
        32: 'Detachable'
    }

    modelos_hardware_aferidor = {
        1: 'Desktop',
        2: 'Desktop',
        3: 'Desktop',
        4: 'Desktop',
        5: 'Desktop',
        6: 'Desktop',
        7: 'Desktop',
        8: 'Notebook',
        9: 'Notebook',
        10: 'Notebook',
        11: 'Notebook',
        12: 'Desktop',
        13: 'Desktop',
        14: 'Notebook',
        15: 'Desktop',
        16: 'Notebook',
        17: 'Desktop',
        18: 'Desktop',
        19: 'Desktop',
        20: 'Desktop',
        21: 'Desktop',
        22: 'Desktop',
        23: 'Desktop',
        24: 'Desktop',
        30: 'Notebook',
        31: 'Notebook',
        32: 'Notebook'
    }

    # Função para escrever uma linha no arquivo de saída
    def write_line(line):
        with open(output_file, "a") as file:
            file.write(line + "\n")

    # Limpar o arquivo de saída (caso já exista)
    if os.path.exists(output_file):
        os.remove(output_file)

    # Função para extrair o valor após o '='
    def extract_value(data):
        return data.split('=')[1].strip()

    # Função para executar um comando e capturar a saída
    def run_command(command):
        result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, shell=True)
        return result.stdout.strip()

    # Função para executar um comando no powershell e capturar a saída
    def run_powershell_command(command):
        result = subprocess.run(['powershell', '-Command', command], stdout=subprocess.PIPE, text=True)
        return result.stdout

    # Lista os softwares obtidos
    def lista_softwares(softwares):
        lista = []

        #Extrair as linhas da resposta
        linhas = softwares.split('\n')

        #Colocar cada linha na lista final
        for linha in linhas:
            lista.append(linha)
        
        return lista


    #------------ ENVIAR ----------------#
    consulta_completa = run_command('wmic path Win32_operatingsystema get Caption, CSName, RegisteredUser, SerialNumber /format:value')
    so_full = f"{run_command('wmic os get Caption /format:value')} ({platform.architecture()[0]})"
    nome_maquina_full = run_command('wmic os get CSName /format:value')

    # Verifica se tem valor e extrai o valor, senão atribui uma mensagem padrão
    so = extract_value(so_full) if so_full else "Sistema Operacional não foi identificado"
    nome_maquina = extract_value(nome_maquina_full) if nome_maquina_full else "O nome da máquina não foi identificado"

    #------------ ENVIAR ----------------#
    #Tipo de dispostivo
    #https://learn.microsoft.com/pt-br/windows/win32/cimwin32prov/win32-systemenclosure

    tipo = run_command('wmic path Win32_SystemEnclosure get chassisTypes /format:value')

    if tipo:

        num_tipo = int((extract_value(tipo)).strip('{}'))
        tipo_hw_wmic_full = modelos_hardware.get(num_tipo)
        tipo_hw_aferidor_full = modelos_hardware_aferidor.get(num_tipo)

        if tipo_hw_aferidor_full and tipo_hw_wmic_full:
            tipo_hw_wmic = f"{tipo_hw_wmic_full} ({num_tipo})"
            tipo_hw_aferidor = f"{tipo_hw_aferidor_full}"

        else:
            tipo_hw_wmic = tipo_hw_aferidor = "Tipo de dispositivo não foi localizado na tabela"
    else:
        tipo_hw_wmic = tipo_hw_aferidor = "Tipo de dispositivo não foi localizado pelo wmi"

    #------------ ENVIAR ----------------#

    # Informações sobre a placa-mãe
    manufacturer_info = run_command('wmic path Win32_BaseBoard get Manufacturer /format:value').strip()
    product_info = run_command('wmic path Win32_BaseBoard get Product /format:value').strip()
    motherboard_info =""

    if manufacturer_info and product_info:
        motherboard_info = f"{extract_value(manufacturer_info)} {extract_value(product_info)}"
    elif manufacturer_info:
        motherboard_info = f"{extract_value(manufacturer_info)}"
    elif product_info:
        motherboard_info = f"{extract_value(product_info)}"

    if not motherboard_info:
        motherboard_info = "Placa mãe não foi localizada!"

    #------------ ENVIAR ----------------#

    # Informações sobre a CPU
    cpu = f"{run_command('wmic path Win32_Processor get Name /format:value')}"
    cpu_arquitetura = platform.architecture()[0]

    if cpu:
        cpu_completa = f"{extract_value(cpu)} ({cpu_arquitetura})"
    else:
        cpu_completa = "CPU não foi localizada!"


    #------------ ENVIAR ----------------#
    # Informações sobre a memória RAM

    #Usando o Win32_ComputerSystem (Melhor metodo)
    ram = run_command('wmic path Win32_ComputerSystem get TotalPhysicalMemory /format:value')

    if ram:
        ram_gb = int(extract_value(ram))/(1024 ** 3)
        ram_gb_aprox = f"{math.ceil(ram_gb)} GB"
    else:
        ram_gb_aprox = "Memória RAM não encontrada!"


    #------------ ENVIAR ----------------#

    # Informações sobre dispositivos de armazenamento
    storage_info = run_command('wmic path Win32_DiskDrive where MediaType="Fixed hard disk media" get caption,model,size,status /format:value')

    #lista
    lista_hds = []

    # Listando
    if storage_info:
        # Dividir a saída em linhas
        lines = storage_info.strip().split('\n')

        # Processar cada linha
        for line in lines:
            parts = line.split('=')
            if len(parts) == 2:
                key, value = parts
                if key == "Caption":
                    caption = value.strip()
                elif key == "Model":
                    model = value.strip()
                elif key == "Size":
                    # Converter o tamanho para GB
                    size_bytes = int(value.strip())
                    size_gb = size_bytes / 1000000000  
                    size_gb_real = size_bytes / (1024**3) # 1 GB = 1024^3 bytes
                elif key == "Status":
                    status = value.strip()
                    hd_info = f"{caption} - {int(size_gb)} GB ({status})"
                    lista_hds.append(hd_info)

    #O metodo join concatena os elementos da lista em uma unica string, separadas pelo '/'
    hds = ' / '.join(lista_hds)

    if not lista_hds:
        hds = "A ferramenta não localizou nenhum dispostivo de armazenamento"
        
    #------------ ENVIAR ----------------#

    # Informações sobre adaptadores de vídeo
    video_info = run_command('wmic path win32_VideoController get name, Description, Caption, Status, VideoProcessor, AdapterRAM /format:value')
    video_lines = video_info.strip().split('\n')
    name = ""
    adapter_ram_mb = ""
    description = ""

    # Loop pelas linhas das informações
    for line in video_lines:
        parts = line.split('=', 1)
        if len(parts) == 2:
            key, value = parts
            if key == 'Name':
                name = value.strip()
            elif key == 'AdapterRAM':
                adapter_ram_bytes = int(value.strip())
                adapter_ram_mb = adapter_ram_bytes / (1024 * 1024)  # Converter para MB
            
            
    # Verifica se todas as informações necessárias foram obtidas
    if name and adapter_ram_mb:
        gpu = f"{name} - ({adapter_ram_mb:.2f} MB)"
    else:
        gpu = "Informações incompletas ou não encontradas."

    # DRIVERS (CD/DVD)

    # Executa o comando WMIC para listar as unidades de CD/DVD
    wmic_command = 'wmic cdrom get Caption /format:value'
    cd_dvd_drives = run_command(wmic_command)

    # Verifica se há unidades de CD/DVD listadas (Se tiver ele coloca o nome no arquivo)
    if cd_dvd_drives:
        cd = extract_value(cd_dvd_drives)
    else:
        cd = "A maquina não tem uma unidade de CD/DVD instalado!"

    # Crie um dicionário com os dados que você deseja enviar
    envio = {
        'data_hora': str(datetime.now()),
        'sistema_operacional': so,
        'nome_maquina': nome_maquina,
        'tipo_maquina_wmi': tipo_hw_wmic,
        'tipo_maquina_aferidor': tipo_hw_aferidor,
        'placa_mae': motherboard_info,
        'cpu': cpu_completa,
        'ram': ram_gb_aprox,
        'hds': hds,
        'gpu': gpu,
        'cd': cd,
        'gpu_completa': video_info
    }
    
    # Exibe e registra os dados nos logs
    write_line(f"------------------- HARDWARE_PC -------------------")
    print(f"------------------- HARDWARE_PC -------------------")


    for chave, valor in envio.items():
        sheet3.append([chave, valor])
        write_line(f"\n{chave} - {valor}")
        print(f"\n{chave} - {valor}")



    # ---------------- INFORMAÇÃO DE SOFTWARE ---------------------------

    #Comando que percorre os 4 registros e filtra apenas os registros que possuem a variavel SystemComponente != 1, ou seja, vai exibir apenas os programas instalados pelo usuario
    #foreach ($UKey in 'HKLM:\SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall\*','HKLM:\SOFTWARE\Wow6432node\Microsoft\Windows\CurrentVersion\Uninstall\*','HKCU:\SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall\*','HKCU:\SOFTWARE\Wow6432node\Microsoft\Windows\CurrentVersion\Uninstall\*'){foreach ($Product in (Get-ItemProperty $UKey -ErrorAction SilentlyContinue)){if($Product.DisplayName -and $Product.SystemComponent -ne 1){$Product.DisplayName}}}

    resumido = run_powershell_command("foreach ($UKey in 'HKLM:\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\*','HKLM:\\SOFTWARE\\Wow6432node\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\*','HKCU:\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\*','HKCU:\\SOFTWARE\\Wow6432node\\Microsoft\\Windows\\CurrentVersion\\Uninstall\*'){foreach ($Product in (Get-ItemProperty $UKey -ErrorAction SilentlyContinue)){if($Product.DisplayName -and $Product.SystemComponent -ne 1){$Product.DisplayName}}}")
    lista_resumido = lista_softwares(resumido)

    completo = run_powershell_command("foreach ($UKey in 'HKLM:\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\*','HKLM:\\SOFTWARE\\Wow6432node\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\*','HKCU:\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\*','HKCU:\\SOFTWARE\\Wow6432node\\Microsoft\\Windows\\CurrentVersion\\Uninstall\*'){foreach ($Product in (Get-ItemProperty $UKey -ErrorAction SilentlyContinue)){if($Product.DisplayName -ne 1){$Product.DisplayName}}}")
    lista_completa = lista_softwares(completo)

    write_line(f"\n\n------------------- SOFTWARE -------------------")
    print(f"\n\n------------------- SOFTWARE -------------------")
    for app in lista_resumido:
        write_line(app)
        print(app)

    for row, valor in enumerate(lista_resumido, start=1):
        sheet.cell(row=row, column=1, value=valor)
    
    for row, valor in enumerate(lista_completa, start=1):
        sheet2.cell(row=row, column=1, value=valor)

    workbook.save(planilha_path)
    workbook.close()

except Exception as e:
    print(f"Ocorreu um erro inesperado! {e}")

print(f"\n\nOs arquivos de log foram salvos em {path}")

input("Aperta qualquer tecla para finalizar...")  

