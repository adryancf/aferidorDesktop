import os
import platform
import subprocess
import math
import requests             
from datetime import datetime

import openpyxl

try:

    # Crie um novo arquivo .xlsx
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Lista Resumida"  
    sheet2 = workbook.create_sheet(title="Lista Completa")
    sheet3 = workbook.create_sheet(title="Hardware")


    '''
    Links Uteis

    https://learn.microsoft.com/en-us/windows/win32/cimwin32prov/computer-system-hardware-classes
    https://learn.microsoft.com/pt-br/windows/win32/cimwin32prov/computer-system-hardware-classes
    https://learn.microsoft.com/pt-br/windows/win32/wmisdk/wmi-tasks--computer-hardware



    '''

    # Dicionário que relaciona os números aos valores correspondentes
    modelos_hardware = {
        1: 'Other',
        2: 'Unknown',
        3: 'Desktop',
        4: 'Low Profile Desktop',
        5: 'Pizza Box',
        6: 'Mini Tower',
        7: 'Tower',
        8: 'Portable',
        9: 'Laptop',
        10: 'Notebook',
        11: 'Hand Held',
        12: 'Docking Station',
        13: 'All in One',
        14: 'Sub Notebook',
        15: 'Space-Saving',
        16: 'Lunch Box',
        17: 'Main System Chassis',
        18: 'Expansion Chassis',
        19: 'SubChassis',
        20: 'Bus Expansion Chassis',
        21: 'Peripheral Chassis',
        22: 'Storage Chassis',
        23: 'Rack Mount Chassis',
        24: 'Sealed-Case PC',
        30: 'Tablet',
        31: 'Convertible',
        32: 'Detachable'
    }

    modelos_hardware_aferidor = {
        1: 'Desktop',
        2: 'Desktop',
        3: 'Desktop',
        4: 'Desktop',
        5: 'Desktop',
        6: 'Desktop',
        7: 'Desktop',
        8: 'Notebook',
        9: 'Notebook',
        10: 'Notebook',
        11: 'Notebook',
        12: 'Desktop',
        13: 'Desktop',
        14: 'Notebook',
        15: 'Desktop',
        16: 'Notebook',
        17: 'Desktop',
        18: 'Desktop',
        19: 'Desktop',
        20: 'Desktop',
        21: 'Desktop',
        22: 'Desktop',
        23: 'Desktop',
        24: 'Desktop',
        30: 'Notebook',
        31: 'Notebook',
        32: 'Notebook'
    }

    # Nome do arquivo de saída
    output_file = "informacoes_sistema.txt"

    # Função para escrever uma linha no arquivo de saída
    def write_line(line):
        with open(output_file, "a") as file:
            file.write(line + "\n")

    # Limpar o arquivo de saída (caso já exista)
    if os.path.exists(output_file):
        os.remove(output_file)

    # Função para extrair o valor após o '='
    def extract_value(data):
        return data.split('=')[1].strip()

    # Função para executar um comando e capturar a saída
    def run_command(command):
        result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, shell=True)
        return result.stdout.strip()

    def run_powershell_command(command):
        result = subprocess.run(['powershell', '-Command', command], stdout=subprocess.PIPE, text=True)
        return result.stdout

    def lista_softwares(softwares):
        lista = []

        #Extrair as linhas da resposta
        linhas = softwares.split('\n')

        #Colocar cada linha na lista final
        for linha in linhas:
            lista.append(linha)
        
        return lista

    def verificar_listas(lista1, lista2):
        if sorted(lista1) == sorted(lista2):
            print ("As listas têm o mesmo conteúdo, independentemente da ordem.")
        else:
            print ("As listas têm conteúdo diferente.")

            # Converta as listas em conjuntos (sets)
            set1 = set(lista1)
            set2 = set(lista2)

            # Encontre os elementos que estão em set1 e não em set2
            elementos_exclusivos = set1 - set2

            # Converta de volta para uma lista (se necessário)
            elementos_exclusivos_lista = list(elementos_exclusivos)

            print("Elementos exclusivos em lista1:", elementos_exclusivos_lista)

    # Informações básicas do sistema
    write_line(f"Data e Hora da Coleta: {datetime.now()}")

    #------------ ENVIAR ----------------#
    consulta_completa = run_command('wmic path Win32_operatingsystema get Caption, CSName, RegisteredUser, SerialNumber /format:value')
    so = f"{extract_value(run_command('wmic os get Caption /format:value'))} ({platform.architecture()[0]})"
    nome_maquina = extract_value(run_command('wmic os get CSName /format:value'))

    write_line(f"\n\nDados Sistema Operacional")

    #Verifica se tem valor
    if so:
        write_line(f"Sistema Operacional = {(so)}")
        if nome_maquina:
            write_line(f"Nome do Sistema = {nome_maquina}")
        else:
            write_line("Nome do Sistema = O nome da maquina nao foi encontrado")
    else:
        write_line("Sistema Operacional não localizado")


    #------------ ENVIAR ----------------#
    #Tipo de dispostivo
    #https://learn.microsoft.com/pt-br/windows/win32/cimwin32prov/win32-systemenclosure

    tipo = extract_value(run_command('wmic path Win32_SystemEnclosure get chassisTypes /format:value'))
    num_tipo = int(tipo.strip('{}'))
    tipo_hw_original = modelos_hardware.get(num_tipo)
    tipo_hw = modelos_hardware_aferidor.get(num_tipo)

    if tipo_hw and tipo_hw_original:
        write_line(f"\nDispostivo Tabela = {tipo_hw_original} ({num_tipo})")
        write_line(f"Dispostivo Aferidor = {tipo_hw}")

    else:
        write_line("\nTipo de dispositivo não foi localizado;/")

    #------------ ENVIAR ----------------#

    # Informações sobre a placa-mãe
    manufacturer_info = run_command('wmic path Win32_BaseBoard get Manufacturer /format:value').strip()
    product_info = run_command('wmic path Win32_BaseBoard get Product /format:value').strip()
    manufacturer = extract_value(manufacturer_info)
    product = extract_value(product_info)

    motherboard_info = f"{manufacturer} {product}"


    write_line("\nPlaca-Mãe: ")
    if motherboard_info:
        write_line(f"{motherboard_info}")
    else:
        write_line("Placa mãe não foi localizada!")

    #------------ ENVIAR ----------------#

    # Informações sobre a CPU
    cpu = f"{extract_value(run_command('wmic path Win32_Processor get Name /format:value'))}"
    cpu_arquitetura = platform.architecture()[0]

    if cpu:
        write_line(f"\nProcessador: {cpu} ({cpu_arquitetura})")
    else:
        write_line("CPU não foi localizada!")
    #------------ ENVIAR ----------------#
    # Informações sobre a memória RAM

    #Usando o Win32_ComputerSystem (Melhor metodo)
    ram_gb = int(extract_value(run_command('wmic path Win32_ComputerSystem get TotalPhysicalMemory /format:value')))/(1024 ** 3)
    ram_gb_aprox = f"{math.ceil(ram_gb)} GB"

    write_line(f"\nWin32_ComputerSystem - Memória RAM:")
    write_line(f"Memória RAM instalada = {ram_gb_aprox} GB")

    #------------ ENVIAR ----------------#

    # Informações sobre dispositivos de armazenamento
    write_line("\nDispositivos de Armazenamento:\n")
    storage_info = run_command('wmic path Win32_DiskDrive where MediaType="Fixed hard disk media" get caption,model,size,status /format:value')

    #lista
    hds_info = []

    # Listando
    if storage_info:
        # Dividir a saída em linhas
        lines = storage_info.strip().split('\n')

        # Processar cada linha
        for line in lines:
            parts = line.split('=')
            if len(parts) == 2:
                key, value = parts
                if key == "Caption":
                    caption = value.strip()
                elif key == "Model":
                    model = value.strip()
                elif key == "Size":
                    # Converter o tamanho para GB
                    size_bytes = int(value.strip())
                    size_gb = size_bytes / 1000000000  
                    size_gb_real = size_bytes / (1024**3) # 1 GB = 1024^3 bytes
                elif key == "Status":
                    status = value.strip()
                    
                    hd_info = f"{caption} - {int(size_gb)} GB ({status})"

                    # Escrever os dados no arquivo de texto
                    write_line(hd_info)
                    hds_info.append(hd_info)

    hds = ' / '.join(hds_info)

    if hds_info:
        write_line(f"\n{hds}")
    else:
        write_line(f"\nA ferramenta não localizou nenhum dispostivo de armazenamento")
        
    #------------ ENVIAR ----------------#

    # Informações sobre adaptadores de vídeo
    write_line("\nAdaptadores de Vídeo:")
    video_info = run_command('wmic path win32_VideoController get name, Description, Caption, Status, VideoProcessor, AdapterRAM /format:value')
    video_lines = video_info.strip().split('\n')
    name = ""
    adapter_ram_mb = ""
    description = ""

    # Loop pelas linhas das informações
    for line in video_lines:
        parts = line.split('=', 1)
        if len(parts) == 2:
            key, value = parts
            if key == 'Name':
                name = value.strip()
            elif key == 'AdapterRAM':
                adapter_ram_bytes = int(value.strip())
                adapter_ram_mb = adapter_ram_bytes / (1024 * 1024)  # Converter para MB
            
            
    # Verifica se todas as informações necessárias foram obtidas
    if name and adapter_ram_mb:
        gpu = f"{name} - ({adapter_ram_mb:.2f} MB)"
        write_line(gpu)
    else:
        write_line("Informações incompletas ou não encontradas.")
        write_line(f"{video_info}")


    # DRIVERS (CD/DVD)

    # Executa o comando WMIC para listar as unidades de CD/DVD
    wmic_command = 'wmic cdrom get Caption /format:value'
    cd_dvd_drives = extract_value(run_command(wmic_command))

    # Verifica se há unidades de CD/DVD listadas (Se tiver ele coloca o nome no arquivo)
    if cd_dvd_drives:

        write_line(f"\n(CD/DVD) - {(cd_dvd_drives)}")
    else:
        write_line("\nA maquina não tem uma unidade de CD/DVD instalado!")

    # Crie um dicionário com os dados que você deseja enviar
    envio = {
        'data_hora_coleta': str(datetime.now()),
        'sistema_operacional': so,
        'nome_maquina': nome_maquina,
        'tipo_maquina_wmi': tipo_hw_original,
        'tipo_maquina': tipo_hw,
        'placa_mae': motherboard_info,
        'cpu': cpu,
        'ram': ram_gb_aprox,
        'hds': hds,
        'gpu': gpu,
        'cd': cd_dvd_drives,
        'gpu_completa': video_info
    }

    """
    for chave, valor in envio.items():
        print(f'{chave}: {valor}')
    """
    # Adicione os dados à planilha
    for chave, valor in envio.items():
        sheet3.append([chave, valor])

    # ---------------- INFORMAÇÃO DE SOFTWARE ---------------------------

    #Comando que percorre os 4 registros e filtra apenas os registros que possuem a variavel SystemComponente != 1, ou seja, vai exibir apenas os programas instalados pelo usuario
    #foreach ($UKey in 'HKLM:\SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall\*','HKLM:\SOFTWARE\Wow6432node\Microsoft\Windows\CurrentVersion\Uninstall\*','HKCU:\SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall\*','HKCU:\SOFTWARE\Wow6432node\Microsoft\Windows\CurrentVersion\Uninstall\*'){foreach ($Product in (Get-ItemProperty $UKey -ErrorAction SilentlyContinue)){if($Product.DisplayName -and $Product.SystemComponent -ne 1){$Product.DisplayName}}}

    resumido = run_powershell_command("foreach ($UKey in 'HKLM:\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\*','HKLM:\\SOFTWARE\\Wow6432node\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\*','HKCU:\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\*','HKCU:\\SOFTWARE\\Wow6432node\\Microsoft\\Windows\\CurrentVersion\\Uninstall\*'){foreach ($Product in (Get-ItemProperty $UKey -ErrorAction SilentlyContinue)){if($Product.DisplayName -and $Product.SystemComponent -ne 1){$Product.DisplayName}}}")
    lista_resumido = lista_softwares(resumido)

    write_line("\n\nLista Resumida:")
    for app in lista_resumido:
        write_line(app)

    for row, valor in enumerate(lista_resumido, start=1):
        sheet.cell(row=row, column=1, value=valor)

    completo = run_powershell_command("foreach ($UKey in 'HKLM:\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\*','HKLM:\\SOFTWARE\\Wow6432node\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\*','HKCU:\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\*','HKCU:\\SOFTWARE\\Wow6432node\\Microsoft\\Windows\\CurrentVersion\\Uninstall\*'){foreach ($Product in (Get-ItemProperty $UKey -ErrorAction SilentlyContinue)){if($Product.DisplayName -ne 1){$Product.DisplayName}}}")
    lista_completa = lista_softwares(completo)

    write_line("\n\nLista Completa:")
    for app in lista_completa:
        write_line(app)

    for row, valor in enumerate(lista_completa, start=1):
        sheet2.cell(row=row, column=1, value=valor)


    # Salve o arquivo .xlsx
    workbook.save("Final.xlsx")

    # Feche o arquivo após salvar
    workbook.close()
    
except Exception as e:
    print(f"Ocorreu um erro inesperado! {e}")

input("Aperta qualquer tecla para finalizar...")  

