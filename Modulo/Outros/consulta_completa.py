import openpyxl

# Crie um novo arquivo .xlsx
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Dados Separados"  
sheet2 = workbook.create_sheet(title="Dados Juntos")

import subprocess

def run_powershell_command(command):
    result = subprocess.run(['powershell', '-Command', command], stdout=subprocess.PIPE, text=True)
    return result.stdout

def lista_softwares(softwares):
    lista = []

    #Extrair as linhas da resposta
    linhas = softwares.split('\n')

    #Colocar cada linha na lista final
    for linha in linhas:
        lista.append(linha)
    
    return lista


#comandos pelo powershell
lista_geral = []

softwares_all_64 = run_powershell_command('Get-ItemProperty HKLM:\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\* | Where-Object { $_.DisplayName -ne $null } | Select-Object -ExpandProperty DisplayName')
if (softwares_all_64.find("Get-ItemProperty")) != 0:
    #erro
    lista_all_64 = lista_softwares(softwares_all_64)
    lista_geral.append(lista_all_64)
else:
    print("ERRO - softwares_all_64")
    lista_all_64 = []


softwares_all_32 = run_powershell_command('Get-ItemProperty HKLM:\\SOFTWARE\\Wow6432Node\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\* | Where-Object { $_.DisplayName -ne $null } | Select-Object -ExpandProperty DisplayName')
if (softwares_all_32.find("Get-ItemProperty")) != 0:
    #erro
    lista_all_32 = lista_softwares(softwares_all_32)
    lista_geral.append(lista_all_32)
else:
    print("ERRO - softwares_all_32")
    lista_all_32 = []



softwares_user_64 = run_powershell_command('Get-ItemProperty HKCU:\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\* | Where-Object { $_.DisplayName -ne $null } | Select-Object -ExpandProperty DisplayName')
if (softwares_user_64.find("Get-ItemProperty")) != 0:
    #erro
    lista_user_64 = lista_softwares(softwares_user_64)
    lista_geral.append(lista_user_64)

else:
    print("ERRO - softwares_user_64")
    lista_user_64 = []

softwares_user_32 = run_powershell_command('Get-ItemProperty HKCU:\\SOFTWARE\\Wow6432Node\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\* | Where-Object { $_.DisplayName -ne $null } | Select-Object -ExpandProperty DisplayName')
if (softwares_user_32.find("Get-ItemProperty")) != 0:
    #erro
    lista_user_32 = lista_softwares(softwares_user_32)
    lista_geral.append(lista_user_32)

else:
    print("ERRO - SOFTWARE_USER_32")
    lista_user_32 = []


lista_separada = [lista_all_32, lista_all_64, lista_user_64, lista_user_32]


# Transponha as listas para que cada lista seja adicionada em uma coluna
for lista in lista_separada:
    for i, valor in enumerate(lista, start=1):
        sheet.cell(row=i, column=lista_separada.index(lista) + 1, value=valor)

dados_combinados = [valor for lista in lista_separada for valor in lista]

for row, valor in enumerate(dados_combinados, start=1):
    sheet2.cell(row=row, column=1, value=valor)